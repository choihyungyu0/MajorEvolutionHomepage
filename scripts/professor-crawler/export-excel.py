#!/usr/bin/env python3
"""Create a collaboration-friendly Excel workbook from normalized professor JSON."""

from __future__ import annotations

import argparse
import json
from collections import defaultdict
from datetime import date
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


SHEET_COLUMNS = {
    "단과대학": ["id", "university", "name"],
    "학과": [
        "id",
        "college_id",
        "name",
        "organization_id",
        "official_homepage_url",
        "discovery_source_url",
        "association_status",
    ],
    "교수": [
        "id",
        "university",
        "name",
        "title",
        "official_profile_url",
        "source_url",
        "status",
        "research_fields_status",
        "publications_status",
        "collected_at",
        "content_hash",
    ],
    "교수_학과": [
        "professor_id",
        "department_id",
        "source_record_id",
        "association_status",
    ],
    "연구분야": ["id", "professor_id", "field"],
    "논문": [
        "id",
        "professor_id",
        "title",
        "publication_type",
        "published_date",
        "date_quality",
        "doi",
        "kci_id",
        "official_profile_url",
        "metadata_source",
    ],
    "논문_메타데이터": [
        "id",
        "publication_id",
        "provider",
        "external_id",
        "canonical_title",
        "journal_title",
        "authors_json",
        "abstract",
        "keywords_json",
        "citation_count",
        "landing_url",
        "license_url",
        "fetched_at",
        "match_method",
        "match_score",
    ],
    "수집이슈": [
        "id",
        "department_id",
        "status",
        "reason",
        "source_url",
        "collected_at",
    ],
}

HEADER_FILL = PatternFill("solid", fgColor="17324D")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SECTION_FILL = PatternFill("solid", fgColor="DDEBF7")
WARNING_FILL = PatternFill("solid", fgColor="FFF2CC")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("normalized_json", type=Path)
    parser.add_argument("output_xlsx", type=Path)
    parser.add_argument("--photo-references", type=Path)
    return parser.parse_args()


def safe_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        value = json.dumps(value, ensure_ascii=False)
    if not isinstance(value, str):
        return value
    value = value[:32767]
    if value.startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value


def date_quality(value: Any, collected_at: str) -> str:
    if not value:
        return "미기재"
    try:
        published = date.fromisoformat(str(value)[:10])
        collected = date.fromisoformat(str(collected_at)[:10])
    except ValueError:
        return "형식 확인 필요"
    return "발행일 확인 필요" if published > collected else "확인"


def add_table_sheet(
    workbook: Workbook,
    title: str,
    rows: Iterable[dict[str, Any]],
    columns: list[str],
) -> None:
    worksheet = workbook.create_sheet(title)
    worksheet.append(columns)
    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row_count = 0
    for row in rows:
        worksheet.append([safe_cell(row.get(column)) for column in columns])
        row_count += 1

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{max(row_count + 1, 1)}"
    worksheet.sheet_view.showGridLines = False
    worksheet.row_dimensions[1].height = 24
    for index, column in enumerate(columns, 1):
        width = min(max(len(column) + 2, 12), 36)
        if "url" in column:
            width = 34
        elif column in {"title", "reason", "abstract", "research_fields"}:
            width = 48
        worksheet.column_dimensions[get_column_letter(index)].width = width
    if row_count:
        table = Table(
            displayName=f"Table{len(workbook.worksheets):02d}",
            ref=f"A1:{get_column_letter(len(columns))}{row_count + 1}",
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)


def build_joined_rows(data: dict[str, Any]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    colleges = {row["id"]: row for row in data["colleges"]}
    departments = {row["id"]: row for row in data["departments"]}
    professor_fields: dict[str, list[str]] = defaultdict(list)
    professor_publications: dict[str, list[dict[str, Any]]] = defaultdict(list)
    professor_departments: dict[str, list[dict[str, Any]]] = defaultdict(list)

    for row in data["research_fields"]:
        professor_fields[row["professor_id"]].append(row["field"])
    for row in data["publications"]:
        professor_publications[row["professor_id"]].append(row)
    for relation in data["professor_departments"]:
        department = departments.get(relation["department_id"], {})
        college = colleges.get(department.get("college_id"), {})
        professor_departments[relation["professor_id"]].append(
            {
                "college": college.get("name", ""),
                "department": department.get("name", ""),
                "association_status": relation.get("association_status", ""),
            }
        )

    joined_professors = []
    joined_publications = []
    professors = {row["id"]: row for row in data["professors"]}
    for professor in data["professors"]:
        affiliations = professor_departments.get(professor["id"], [{}])
        publications = professor_publications.get(professor["id"], [])
        published_dates = [
            str(item["published_date"])
            for item in publications
            if item.get("published_date")
        ]
        joined_professors.append(
            {
                "professor_id": professor["id"],
                "name": professor["name"],
                "title": professor.get("title"),
                "university": professor["university"],
                "colleges": " · ".join(
                    sorted({item.get("college", "") for item in affiliations if item.get("college")})
                ),
                "departments": " · ".join(
                    sorted(
                        {
                            item.get("department", "")
                            for item in affiliations
                            if item.get("department")
                        }
                    )
                ),
                "association_status": " · ".join(
                    sorted(
                        {
                            item.get("association_status", "")
                            for item in affiliations
                            if item.get("association_status")
                        }
                    )
                ),
                "research_fields": " · ".join(professor_fields.get(professor["id"], [])),
                "publication_count": len(publications),
                "latest_published_date": max(published_dates, default=""),
                "status": professor["status"],
                "research_fields_status": professor["research_fields_status"],
                "publications_status": professor["publications_status"],
                "official_profile_url": professor.get("official_profile_url"),
                "collected_at": professor["collected_at"],
            }
        )

    for publication in data["publications"]:
        professor = professors.get(publication["professor_id"], {})
        affiliations = professor_departments.get(publication["professor_id"], [])
        joined_publications.append(
            {
                "publication_id": publication["id"],
                "professor_id": publication["professor_id"],
                "professor_name": professor.get("name", ""),
                "departments": " · ".join(
                    sorted(
                        {
                            item.get("department", "")
                            for item in affiliations
                            if item.get("department")
                        }
                    )
                ),
                "title": publication["title"],
                "publication_type": publication.get("publication_type"),
                "published_date": publication.get("published_date"),
                "date_quality": date_quality(
                    publication.get("published_date"),
                    professor.get("collected_at", data["generated_at"]),
                ),
                "doi": publication.get("doi"),
                "kci_id": publication.get("kci_id"),
                "official_profile_url": publication.get("official_profile_url"),
                "metadata_source": publication.get("metadata_source"),
            }
        )
    return joined_professors, joined_publications


def add_overview(workbook: Workbook, data: dict[str, Any]) -> None:
    worksheet = workbook.active
    worksheet.title = "사용안내"
    worksheet.sheet_view.showGridLines = False
    worksheet["A1"] = "단국대학교 교수·연구 데이터"
    worksheet["A1"].font = Font(size=18, bold=True, color="17324D")
    worksheet.merge_cells("A1:D1")
    worksheet["A3"] = "기준"
    worksheet["A3"].fill = SECTION_FILL
    worksheet["A3"].font = Font(bold=True)
    details = [
        ("생성일", data.get("generated_at", "")),
        ("대학", data.get("university", "")),
        ("논문 범위", "대학 공식 교수 프로필에 노출된 연구실적만"),
        ("주의", "공식 프로필 미기재는 연구·논문이 없다는 뜻이 아닙니다."),
        ("사진", "사진 파일은 포함하지 않으며 공식 프로필 링크를 사용합니다."),
    ]
    for row_index, (label, value) in enumerate(details, 4):
        worksheet.cell(row_index, 1, label).font = Font(bold=True)
        worksheet.cell(row_index, 2, safe_cell(value))

    worksheet["A10"] = "데이터 규모"
    worksheet["A10"].fill = SECTION_FILL
    worksheet["A10"].font = Font(bold=True)
    count_labels = {
        "colleges": "단과대학",
        "departments": "학과·전공",
        "professors": "중복 제거 교수",
        "professor_department_links": "교수-학과 연결",
        "research_fields": "연구분야",
        "publications": "공식 프로필 연구실적",
        "publication_metadata": "외부 메타데이터",
        "collection_issues": "수집 이슈",
    }
    for row_index, (key, label) in enumerate(count_labels.items(), 11):
        worksheet.cell(row_index, 1, label)
        worksheet.cell(row_index, 2, data.get("counts", {}).get(key, 0))

    worksheet["A21"] = "품질 해석"
    worksheet["A21"].fill = WARNING_FILL
    worksheet["A21"].font = Font(bold=True)
    notes = [
        "FOUND: 공식 프로필에서 연구분야 또는 연구실적을 확인했습니다.",
        "NOT_LISTED_ON_OFFICIAL_PROFILE: 공식 프로필에 해당 목록이 표시되지 않았습니다.",
        "PROFILE_UNAVAILABLE: 학과 공식 교수 페이지를 확인하지 못했습니다.",
        "발행일 확인 필요: 수집일보다 미래 날짜가 공식 프로필에 표시된 값입니다.",
        "SHARED_OFFICIAL_HOMEPAGE: 여러 전공이 같은 공식 홈페이지를 공유합니다.",
    ]
    for row_index, note in enumerate(notes, 22):
        worksheet.cell(row_index, 1, "•")
        worksheet.cell(row_index, 2, note)
    worksheet.column_dimensions["A"].width = 24
    worksheet.column_dimensions["B"].width = 90
    worksheet.freeze_panes = "A3"


def main() -> None:
    args = parse_args()
    data = json.loads(args.normalized_json.read_text(encoding="utf-8"))
    args.output_xlsx.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    add_overview(workbook, data)
    joined_professors, joined_publications = build_joined_rows(data)
    add_table_sheet(
        workbook,
        "교수_통합",
        joined_professors,
        [
            "professor_id",
            "name",
            "title",
            "university",
            "colleges",
            "departments",
            "association_status",
            "research_fields",
            "publication_count",
            "latest_published_date",
            "status",
            "research_fields_status",
            "publications_status",
            "official_profile_url",
            "collected_at",
        ],
    )
    add_table_sheet(
        workbook,
        "논문_통합",
        joined_publications,
        [
            "publication_id",
            "professor_id",
            "professor_name",
            "departments",
            "title",
            "publication_type",
            "published_date",
            "date_quality",
            "doi",
            "kci_id",
            "official_profile_url",
            "metadata_source",
        ],
    )

    raw_mappings = [
        ("단과대학", "colleges"),
        ("학과", "departments"),
        ("교수", "professors"),
        ("교수_학과", "professor_departments"),
        ("연구분야", "research_fields"),
        ("논문", "publications"),
        ("논문_메타데이터", "publication_metadata"),
        ("수집이슈", "collection_issues"),
    ]
    for sheet_name, key in raw_mappings:
        rows = data.get(key, [])
        if sheet_name == "논문":
            professors = {row["id"]: row for row in data["professors"]}
            rows = [
                {
                    **row,
                    "date_quality": date_quality(
                        row.get("published_date"),
                        professors.get(row["professor_id"], {}).get(
                            "collected_at", data["generated_at"]
                        ),
                    ),
                }
                for row in rows
            ]
        add_table_sheet(workbook, sheet_name, rows, SHEET_COLUMNS[sheet_name])

    if args.photo_references:
        photo_data = json.loads(args.photo_references.read_text(encoding="utf-8"))
        photo_rows = photo_data.get("references", [])
        add_table_sheet(
            workbook,
            "사진_참조",
            photo_rows,
            [
                "professor_id",
                "professor_name",
                "official_profile_url",
                "photo_source_url",
                "source_page_url",
                "checked_at",
                "usage_status",
            ],
        )

    workbook.save(args.output_xlsx)
    print(
        json.dumps(
            {
                "output": str(args.output_xlsx.resolve()),
                "professors": len(joined_professors),
                "publications": len(joined_publications),
                "sheets": workbook.sheetnames,
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
